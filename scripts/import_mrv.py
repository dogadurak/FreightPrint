"""Turn a THETIS-MRV annual report into a factor this engine can be checked against.

The EU publishes, for every ship over 5,000 GT calling at an EEA port, its **verified**
annual CO2, fuel burn, distance sailed and transport work. That is the only independent
observation of the number this project's headline rests on: GLEC's ro-ro figure of
0.068 kg CO2 per tonne-kilometre, which nothing here had ever been able to check.

It cannot be fetched automatically. The portal is a JavaScript application, the download
sits behind a reCAPTCHA, and EMSA publishes no direct file URL — so this script takes
the file after a person has clicked once:

    1. https://mrv.emsa.europa.eu/#public/emission-report
    2. Choose a reporting period and export the table (Excel)
    3. Save it under data/external/, then:

       python scripts/import_mrv.py "data/external/<file>.xlsx" --describe
       python scripts/import_mrv.py "data/external/<file>.xlsx"

`--describe` first, always. EMSA changes these workbooks between reporting periods —
the published file has two rows of grouping labels above the real header and a hundred
and thirteen columns, several of which differ only by the words "on laden voyages" —
and guessing at a column is exactly the shortcut this project does not take. Nothing is
written until every needed column has been found and named in the output.
"""

import argparse
import csv
import statistics
import sys
import warnings
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
OUT = REPO / "data" / "external" / "roro_intensity_mrv.csv"

# GLEC Framework 2019 (Jul 2022) Table 45, ro-ro fleet average, trailer only.
#
# **The tank-to-wheel figure, not the well-to-wheel one, and the difference is the whole
# comparison.** MRV reports the CO2 a ship actually emitted from the fuel it burned —
# that is TTW by construction. GLEC's WTW value of 0.068 additionally carries the
# emissions of producing and delivering that fuel, which no ship reports and no verifier
# checks. Holding 0.068 against an MRV figure would charge the observation for something
# it never measured and make the factor look 8% worse than it is.
GLEC_RORO_TTW = 0.063
GLEC_RORO_WTW = 0.068

# One nautical mile in kilometres. MRV reports per tonne-nautical-mile and GLEC per
# tonne-kilometre; comparing them unconverted is wrong by this factor.
KM_PER_NAUTICAL_MILE = 1.852

# What the derivation needs, and how each column is recognised. Matched on a lower-cased
# substring so a changed unit suffix does not break it, with `not_` to keep a
# near-identical neighbour out: the workbook publishes "per transport work (mass)" twice,
# once over all voyages and once over laden voyages only, and they are different numbers.
WANTED = {
    "imo": {"any": ["imo number"]},
    "ship_type": {"any": ["ship type"]},
    "reporting_period": {"any": ["reporting period"]},
    # The comparable figure: CO2 per tonne-nautical-mile over all voyages, which is the
    # basis a fleet-average factor describes.
    "co2_per_tw": {
        "any": ["emissions per transport work (mass)"],
        "not_": ["laden", "co₂eq", "co2eq", "fuel consumption"],
    },
    # The same over laden voyages only. Kept because the gap between the two is the
    # ballast share — the sea's own version of empty running.
    "co2_per_tw_laden": {
        "any": ["emissions per transport work (mass) on laden"],
        "not_": ["co₂eq", "co2eq", "fuel consumption"],
    },
    # Freight-only allocation, published for ships that also carry passengers. GLEC's
    # Table 45 is trailer-only, so for a ro-pax this is the honest comparison.
    "co2_per_tw_freight": {
        "any": ["emissions per transport work (freight)"],
        "not_": ["laden", "co₂eq", "co2eq", "fuel consumption"],
    },
}

REQUIRED = {"imo", "ship_type", "co2_per_tw"}

# The header is not the first row: two rows of grouping labels sit above it, so reading
# row 1 finds nothing at all — which is exactly what the guard reported the first time
# this ran against a real export.
HEADER_MARKER = "imo number"
MAX_HEADER_SCAN = 12

# The ship types this corridor actually uses. A vehicle carrier moves cars and a
# passenger ship moves people; holding GLEC's trailer factor against either would be
# comparing different vessels and calling it validation.
RORO_TYPES = ("ro-ro ship", "ro-pax ship", "container/ro-ro cargo ship")

# MRV writes this where a ship reported no transport work. Read as a number it is
# nothing, and nothing would drag the median to the floor.
NOT_A_NUMBER = ("division by zero!", "not applicable", "n/a", "")


def find_columns(header: list) -> tuple[dict[str, int], list[str]]:
    """Map each wanted field onto a column, and say which were not found."""
    lowered = [str(cell or "").strip().lower() for cell in header]
    found: dict[str, int] = {}
    for field, rule in WANTED.items():
        for spelling in rule["any"]:
            match = next(
                (
                    i for i, cell in enumerate(lowered)
                    if spelling in cell
                    and not any(bad in cell for bad in rule.get("not_", []))
                ),
                None,
            )
            if match is not None:
                found[field] = match
                break
    return found, [field for field in WANTED if field not in found]


def header_row(sheet) -> tuple[int, list]:
    """Find the row that actually names the columns, rather than assuming the first."""
    for number, row in enumerate(sheet.iter_rows(max_row=MAX_HEADER_SCAN, values_only=True), 1):
        if any(HEADER_MARKER in str(cell or "").lower() for cell in row):
            return number, list(row)
    raise LookupError(
        f"ilk {MAX_HEADER_SCAN} satirda '{HEADER_MARKER}' iceren bir baslik satiri yok"
    )


def _number(value) -> float | None:
    """A reported figure, or None where the ship reported none."""
    if value is None or str(value).strip().lower() in NOT_A_NUMBER:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if number > 0 else None


def describe(path: Path) -> None:
    """Say what is in the workbook without writing anything."""
    import openpyxl

    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for name in book.sheetnames:
            sheet = book[name]
            print(f"\n--- {name}  ({sheet.max_row} satir x {sheet.max_column} sutun) ---")
            try:
                number, header = header_row(sheet)
            except LookupError as error:
                print(f"    {error}")
                continue
            print(f"    baslik satiri: {number}")
            found, missing = find_columns(header)
            for field, index in found.items():
                print(f"    {field:22} <- [{index:3}] {str(header[index])[:56]}")
            for field in missing:
                mark = "ZORUNLU" if field in REQUIRED else "istege bagli"
                print(f"    {field:22} BULUNAMADI ({mark})")
    finally:
        book.close()


def derive(path: Path, sheet_name: str | None = None) -> int:
    """Write one row per ro-ro ship: its verified transport-work intensity."""
    import openpyxl

    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        # The full reports, not the partial ones: a partial return covers part of a year
        # and its intensity is not comparable with an annual figure.
        name = sheet_name or next(
            (s for s in book.sheetnames if "full" in s.lower()), book.sheetnames[0]
        )
        sheet = book[name]
        start, header = header_row(sheet)
        found, _ = find_columns(header)

        absent = REQUIRED - set(found)
        if absent:
            print(
                f"HATA: zorunlu sutun(lar) bulunamadi: {', '.join(sorted(absent))}.\n"
                f"Once --describe ile calistirin; basliklar tahmin edilmeyecek.",
                file=sys.stderr,
            )
            return 0

        out_rows = []
        for row in sheet.iter_rows(min_row=start + 1, values_only=True):
            ship_type = str(row[found["ship_type"]] or "").strip()
            if ship_type.lower() not in RORO_TYPES:
                continue
            per_tw = _number(row[found["co2_per_tw"]])
            if per_tw is None:
                continue

            laden = _number(row[found["co2_per_tw_laden"]]) if "co2_per_tw_laden" in found else None
            freight = (
                _number(row[found["co2_per_tw_freight"]]) if "co2_per_tw_freight" in found else None
            )
            to_kg_per_tonne_km = lambda g: round(g / KM_PER_NAUTICAL_MILE / 1000, 6)

            out_rows.append({
                "imo": row[found["imo"]],
                "ship_type": ship_type,
                "reporting_period": (
                    int(float(row[found["reporting_period"]]))
                    if "reporting_period" in found and row[found["reporting_period"]]
                    else ""
                ),
                "g_co2_per_tonne_nm": round(per_tw, 4),
                "kg_co2_per_tonne_km": to_kg_per_tonne_km(per_tw),
                "kg_co2_per_tonne_km_laden": to_kg_per_tonne_km(laden) if laden else "",
                "kg_co2_per_tonne_km_freight": to_kg_per_tonne_km(freight) if freight else "",
            })

        if not out_rows:
            print(f"HATA: '{name}' sayfasinda ro-ro tipinde gemi bulunamadi.", file=sys.stderr)
            return 0

        # Merge rather than overwrite, so importing each year in turn builds one file.
        # Rows for the period being imported are replaced, which makes a re-run after a
        # corrected publication idempotent instead of doubling the fleet.
        periods = {row["reporting_period"] for row in out_rows}
        kept = []
        if OUT.exists():
            with OUT.open(encoding="utf-8") as f:
                kept = [
                    row for row in csv.DictReader(f)
                    if row.get("reporting_period") not in {str(p) for p in periods}
                ]

        OUT.parent.mkdir(parents=True, exist_ok=True)
        merged = kept + [{k: str(v) for k, v in row.items()} for row in out_rows]
        merged.sort(key=lambda row: (row["reporting_period"], row["imo"]))
        with OUT.open("w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=list(out_rows[0]))
            writer.writeheader()
            writer.writerows(merged)

        values = sorted(r["kg_co2_per_tonne_km"] for r in out_rows)
        median = statistics.median(values)
        print(f"'{name}' -> {len(out_rows)} ro-ro gemisi ({len(merged)} toplam) -> {OUT}")
        print(f"  medyan   {median:.4f} kg CO2/ton-km")
        print(f"  ceyrekler {values[len(values)//4]:.4f} / {values[3*len(values)//4]:.4f}")
        print(f"  aralik   {values[0]:.4f} - {values[-1]:.4f}")
        print(f"  GLEC Tablo 45 (TTW): {GLEC_RORO_TTW}  ->  medyanin {GLEC_RORO_TTW/median:.2f} kati")
        inside = sum(1 for v in values if v <= GLEC_RORO_TTW) / len(values)
        print(f"  gozlenen gemilerin %{inside*100:.0f}'i GLEC'in altinda")
        print(f"  (WTW {GLEC_RORO_WTW} ile karsilastirilmaz: MRV yakit uretimini olcmez)")
        return len(out_rows)
    finally:
        book.close()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    parser.add_argument("workbook", nargs="?", type=Path, help="THETIS-MRV export (.xlsx)")
    parser.add_argument("--describe", action="store_true",
                        help="only report the workbook's columns, write nothing")
    parser.add_argument("--sheet", default=None, help="which worksheet to read")
    args = parser.parse_args()

    if args.workbook is None:
        print(__doc__)
        return 1
    if not args.workbook.exists():
        print(f"dosya yok: {args.workbook}", file=sys.stderr)
        return 1

    # openpyxl complains about the publication's missing default style on every open.
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    if args.describe:
        describe(args.workbook)
        return 0
    return 0 if derive(args.workbook, args.sheet) else 1


if __name__ == "__main__":
    sys.exit(main())
