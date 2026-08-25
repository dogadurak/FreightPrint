"""The customer-facing formats. What matters here is that the basis survives the trip.

A spreadsheet gets sorted and pasted from, a PDF gets attached to somebody else's
filing, and in both cases a carbon figure that has lost the factor set it was produced
on cannot be checked by whoever ends up holding it.
"""

import io

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader

from app.core import route as route_module
from app.core.deliverable import report_to_pdf, report_to_xlsx
from app.core.network import haversine_km
from app.core.report import build_report, parse_shipments
from app.core.road import RoadRoute, RoadRoutingError

# Turkish place names on purpose: the dotless i, s-cedilla and g-breve are absent from
# Latin-1, which is what reportlab's built-in fonts use.
TURKISH_CSV = (
    "reference,origin_name,origin_lon,origin_lat,destination_name,"
    "destination_lon,destination_lat,tonnage\n"
    "SEV-1,Gebze Şubesi,29.4306,40.7889,Düsseldorf,6.7735,51.2277,24\n"
    "SEV-2,Iğdır Deposu,29.43,40.78,Viyana,16.37,48.21,18\n"
)


@pytest.fixture(autouse=True)
def offline_road(monkeypatch):
    def fake(origin, destination):
        km = haversine_km(origin, destination) * 1.3
        return RoadRoute(distance_km=km, duration_h=km / 70, geometry=(origin, destination))

    monkeypatch.setattr(route_module, "road_route", fake)


@pytest.fixture
def report():
    return build_report(parse_shipments(TURKISH_CSV), factor_set="glec", scope="WTW")


def pdf_text(data: bytes) -> str:
    return " ".join(
        " ".join((page.extract_text() or "").split()) for page in PdfReader(io.BytesIO(data)).pages
    )


def sheet_text(workbook, name: str) -> str:
    return " ".join(
        str(cell.value) for row in workbook[name].iter_rows() for cell in row if cell.value
    )


def test_the_workbook_carries_its_basis_on_a_sheet_of_its_own(report):
    """Not a banner above the header row: that is the first thing lost when somebody
    sorts or copies the data out."""
    workbook = load_workbook(io.BytesIO(report_to_xlsx(report)))

    assert workbook.sheetnames == ["Rapor", "Esas ve kaynaklar"]
    basis = sheet_text(workbook, "Esas ve kaynaklar")
    assert "glec" in basis and "WTW" in basis
    assert "GLEC Framework" in basis
    assert "Yöntem" in basis


def test_the_data_sheet_holds_every_shipment_with_a_frozen_header(report):
    workbook = load_workbook(io.BytesIO(report_to_xlsx(report)))
    sheet = workbook["Rapor"]

    assert sheet.max_row == len(report.rows) + 1
    assert sheet.freeze_panes == "A2"
    assert sheet["A1"].value == "Referans"


def test_turkish_names_survive_the_spreadsheet(report):
    workbook = load_workbook(io.BytesIO(report_to_xlsx(report)))

    text = sheet_text(workbook, "Rapor")
    assert "Gebze Şubesi" in text and "Iğdır Deposu" in text


def test_turkish_survives_the_pdf(report):
    """reportlab's built-in fonts are Latin-1, which has no ı, ş, ğ or İ. A Turkish
    report set in them comes out broken, so the bundled Vera face is registered instead."""
    text = pdf_text(report_to_pdf(report))

    for word in ("Gebze Şubesi", "Iğdır Deposu", "Kalkış", "Varış", "Hesap esası", "Yöntem"):
        assert word in text, f"{word!r} did not survive the PDF"


def test_the_pdf_states_the_basis_it_was_produced_on(report):
    text = pdf_text(report_to_pdf(report))

    assert "glec" in text and "WTW" in text
    assert "GLEC Framework" in text


def test_both_formats_are_what_they_claim_to_be(report):
    assert report_to_xlsx(report)[:2] == b"PK"       # xlsx is a zip
    assert report_to_pdf(report)[:5] == b"%PDF-"


def test_the_all_road_baseline_travels_beside_the_chosen_total(report):
    """The report carries the lowest-emission option, so its difference is never
    negative — where multimodal loses, all-road simply is the chosen route and the
    difference is zero. That makes the baseline column the only way a reader can see
    what the alternative would have cost, so it has to be in both formats."""
    assert all(row.saving_co2_kg >= 0 for row in report.calculated), (
        "the report is meant to report the best option"
    )

    workbook = load_workbook(io.BytesIO(report_to_xlsx(report)))
    header = [cell.value for cell in workbook["Rapor"][1]]
    assert "Tam karayolu kg" in header and "Toplam kg CO2" in header

    text = pdf_text(report_to_pdf(report))
    assert "Tam karayolu kg" in text and "Toplam kg CO2" in text


def test_the_method_note_explains_what_a_zero_difference_means(report):
    """Zero is the interesting case on this corridor and reads as "not calculated"
    unless the document says otherwise."""
    text = pdf_text(report_to_pdf(report))

    assert "kazandırmadığı" in text


def test_an_unverified_factor_warns_in_both_formats():
    """HVO and electric are derived, not published. A report priced on one has to say so
    wherever it is read, not only in the CSV."""
    report = build_report(
        parse_shipments(TURKISH_CSV), factor_set="glec", scope="WTW", road_fuel_type="hvo"
    )
    assert report.warnings, "the fixture no longer exercises an unverified factor"

    assert "Uyarılar" in pdf_text(report_to_pdf(report))
    workbook = load_workbook(io.BytesIO(report_to_xlsx(report)))
    assert "Uyarılar" in sheet_text(workbook, "Esas ve kaynaklar")


def test_a_failed_shipment_is_listed_with_its_reason(monkeypatch):
    calls = {"n": 0}
    original = route_module.road_route

    def sometimes_failing(origin, destination):
        calls["n"] += 1
        if calls["n"] == 1:
            raise RoadRoutingError("no road access")
        return original(origin, destination)

    monkeypatch.setattr(route_module, "road_route", sometimes_failing)
    report = build_report(parse_shipments(TURKISH_CSV))

    assert report.failed
    workbook = load_workbook(io.BytesIO(report_to_xlsx(report)))
    assert "no road access" in sheet_text(workbook, "Rapor")
    assert "no road access" in pdf_text(report_to_pdf(report))


def test_an_empty_report_still_produces_a_readable_file():
    """A file of shipments that all failed still has to hand back something that opens
    and says why, rather than an exception the user cannot act on."""
    report = build_report([])

    assert load_workbook(io.BytesIO(report_to_xlsx(report))).sheetnames
    assert "Hesap esası" in pdf_text(report_to_pdf(report))


@pytest.mark.parametrize(
    "value,decimals,expected",
    [(4770, 0, "4.770"), (1234567, 0, "1.234.567"), (24, 0, "24"), (4.5, 1, "4,5"), (0, 0, "0")],
)
def test_numbers_are_written_in_turkish(value, decimals, expected):
    """Not cosmetic. Turkish groups thousands with a full stop, so the English default
    renders 4770 kg as "4,770" — which reads as four point seven seven. In a carbon
    report that is a factor of a thousand, in the direction of looking harmless."""
    from app.core.deliverable import _number_tr

    assert _number_tr(value, decimals) == expected


def test_the_pdf_does_not_print_english_separators(report):
    """The whole document is Turkish; a number in another convention inside it is a
    misreading waiting to happen."""
    text = pdf_text(report_to_pdf(report))

    import re
    english_grouped = re.findall(r"\d,\d{3}\b", text)
    assert not english_grouped, f"English thousands grouping in a Turkish report: {english_grouped}"


def test_whole_tonnages_lose_the_decimal_excel_gave_them(report):
    """Every spreadsheet number arrives as a float, so 24 tonnes becomes 24.0."""
    text = pdf_text(report_to_pdf(report))

    assert "24,0" not in text and "24.0" not in text


def test_the_status_column_is_in_the_documents_own_language(report):
    assert "hesaplandı" in pdf_text(report_to_pdf(report))


def test_every_printed_column_has_a_heading_and_a_value(report):
    """The three lists that describe a column have to agree, and nothing makes them.

    `OUTPUT_COLUMNS` names what the workbook prints, `PDF_COLUMNS` what the PDF prints,
    `HEADER_TR` what each is called and `_row_values` what goes under it. A column added
    to one and not the others is a `KeyError` at render time, which is how both export
    formats came to fail at once: `carrier` reached OUTPUT_COLUMNS alone.
    """
    from app.core.deliverable import HEADER_TR, PDF_COLUMNS, _row_values
    from app.core.report import OUTPUT_COLUMNS

    named = set(HEADER_TR)
    printed = set(OUTPUT_COLUMNS) | set(PDF_COLUMNS)
    assert printed <= named, f"printed with no heading: {sorted(printed - named)}"

    produced = set(_row_values(report.rows[0]))
    assert printed <= produced, f"printed with no value: {sorted(printed - produced)}"


def test_the_pdf_prints_a_subset_of_what_the_workbook_does():
    """The PDF is the narrow view of the same report, not a different report."""
    from app.core.deliverable import PDF_COLUMNS
    from app.core.report import OUTPUT_COLUMNS

    assert set(PDF_COLUMNS) <= set(OUTPUT_COLUMNS)
