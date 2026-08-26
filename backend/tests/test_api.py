import time

import pytest
from fastapi.testclient import TestClient

from app.core import route as route_module
from app.core.road import RoadRoute, RoadRoutingError
from app.core.network import haversine_km
from app.main import app

ROAD_DETOUR_FACTOR = 1.3
AVERAGE_ROAD_SPEED_KMH = 70.0

GEBZE = {"lon": 29.4306, "lat": 40.7889}
DUSSELDORF = {"lon": 6.7735, "lat": 51.2277}


@pytest.fixture
def client(monkeypatch):
    """Serve routes without touching OSRM, so the API contract is what is under test."""

    def fake_road_route(origin, destination):
        distance_km = haversine_km(origin, destination) * ROAD_DETOUR_FACTOR
        return RoadRoute(
            distance_km=distance_km,
            duration_h=distance_km / AVERAGE_ROAD_SPEED_KMH,
            geometry=(origin, destination),
        )

    monkeypatch.setattr(route_module, "road_route", fake_road_route)
    return TestClient(app)


def _post(client, **overrides):
    body = {"origin": GEBZE, "destination": DUSSELDORF} | overrides
    return client.post("/api/routes", json=body)


def test_health(client):
    assert client.get("/health").json() == {"status": "ok"}


def test_terminals_flag_the_ones_no_service_reaches(client):
    terminals = client.get("/api/terminals").json()

    assert len(terminals) == 16
    assert [t["name"] for t in terminals if not t["is_connected"]] == ["Ambarli"]


def test_factor_sets_expose_the_sea_basis_behind_each_choice(client):
    sets = {s["name"]: s for s in client.get("/api/factor-sets").json()}

    assert sets["glec"]["sea_factor_by_scope"]["TTW"] == 0.063
    assert sets["glec_accompanied"]["sea_factor_by_scope"]["TTW"] == 0.093
    assert sets["glec_freight_average"]["sea_factor_by_scope"]["TTW"] == 0.042
    assert sets["placeholder"]["is_verified"] is False


def test_routes_return_the_baseline_first_then_cleanest(client):
    payload = _post(client).json()
    alternatives = payload["alternatives"]

    assert alternatives[0]["is_all_road"]
    others = [a["total_co2_kg"] for a in alternatives[1:]]
    assert others == sorted(others)


def test_response_names_the_factor_set_and_its_sources(client):
    payload = _post(client, factor_set="glec", scope="WTW").json()

    assert payload["factor_set"] == "glec"
    assert payload["scope"] == "WTW"
    assert all("GLEC" in source for source in payload["sources"])


def test_a_negative_saving_is_reported_rather_than_hidden(client):
    """On a ro-ro corridor the multimodal option can lose, and the API must say so."""
    payload = _post(client, factor_set="glec", scope="WTW").json()
    multimodal = [a for a in payload["alternatives"] if not a["is_all_road"]]

    assert multimodal
    assert any(a["saving_co2_kg"] < 0 for a in multimodal)
    for alternative in multimodal:
        if alternative["saving_co2_kg"] < 0:
            assert set(alternative["tree_equivalent"].values()) == {0}


def test_road_and_sea_legs_carry_geometry_but_rail_does_not(client):
    """Sea tracks come from searoute; rail has no computed geometry, so it stays schematic."""
    payload = _post(client).json()
    legs = [leg for a in payload["alternatives"] for leg in a["legs"]]

    assert any(leg["mode"] == "road" and leg["geometry"] for leg in legs)
    assert any(leg["mode"] == "sea" and leg["geometry"] for leg in legs)
    assert all(not leg["geometry"] for leg in legs if leg["mode"] == "rail")


def test_max_alternatives_trims_after_ranking_by_emissions(client):
    full = _post(client, factor_set="glec").json()["alternatives"]
    trimmed = _post(client, factor_set="glec", max_alternatives=1).json()["alternatives"]

    assert len(trimmed) == 2
    assert trimmed[1]["label"] == full[1]["label"]


def test_a_fuel_missing_from_the_set_is_a_422_not_another_set(client):
    response = _post(client, factor_set="glec", road_fuel_type="diesel")

    assert response.status_code == 422
    assert "diesel_b5" in response.json()["detail"]


@pytest.mark.parametrize(
    "overrides",
    [
        {"origin": {"lon": 999, "lat": 40}},
        {"tonnage": 0},
        {"scope": "CRADLE"},
        {"load_factor": 0},
        {"empty_return_share": 5},
    ],
)
def test_impossible_input_is_refused(client, overrides):
    assert _post(client, **overrides).status_code == 422


def test_unreachable_origin_answers_422_not_500(client, monkeypatch):
    def unreachable(origin, destination):
        raise RoadRoutingError("(-21.9, 64.1) is 762 km from the nearest road; no road access")

    monkeypatch.setattr(route_module, "road_route", unreachable)

    response = _post(client)
    assert response.status_code == 422
    assert "no road access" in response.json()["detail"]


def test_the_page_and_its_assets_are_served(client):
    assert client.get("/").status_code == 200
    assert client.get("/static/app.js").status_code == 200
    assert client.get("/static/style.css").status_code == 200


SHIPMENT_CSV = (
    "reference,origin_lon,origin_lat,destination_lon,destination_lat,tonnage\n"
    "SEV-1,29.4306,40.7889,6.7735,51.2277,24\n"
    "SEV-2,29.4306,40.7889,16.3738,48.2082,18\n"
)


def _upload(client, content=SHIPMENT_CSV, **fields):
    return client.post(
        "/api/report",
        files={"file": ("shipments.csv", content, "text/csv")},
        data={"scope": "TTW", "factor_set": "reference"} | fields,
    )


def test_report_returns_a_downloadable_csv(client):
    response = _upload(client)

    assert response.status_code == 200
    assert "text/csv" in response.headers["content-type"]
    assert "attachment" in response.headers["content-disposition"]
    assert response.text.count("SEV-") == 2


def test_report_states_the_factor_set_it_priced_with(client):
    response = _upload(client, factor_set="glec", scope="WTW")

    assert "factor set: glec" in response.text
    assert "GLEC Framework" in response.text


def test_report_rejects_a_file_missing_required_columns(client):
    response = _upload(client, content="origin_lon,origin_lat\n29.43,40.78\n")

    assert response.status_code == 422
    assert "missing column" in response.json()["detail"]


def test_report_rejects_a_factor_set_that_cannot_price_any_row(client):
    response = _upload(client, factor_set="reference", scope="WTW")

    assert response.status_code == 422
    assert "no WTW factor" in response.json()["detail"]


def test_report_rejects_a_non_utf8_upload(client):
    response = client.post(
        "/api/report",
        files={"file": ("shipments.csv", b"\xff\xfe\x00bad", "text/csv")},
        data={"scope": "TTW", "factor_set": "reference"},
    )

    assert response.status_code == 422


def test_the_sample_shipment_file_is_served_and_parses(client):
    response = client.get("/static/ornek_sevkiyatlar.csv")

    assert response.status_code == 200
    assert _upload(client, content=response.text).status_code == 200


SCENARIOS = [
    {"factor_set": "reference", "scope": "TTW"},
    {"factor_set": "glec", "scope": "TTW"},
    {"factor_set": "glec", "scope": "WTW"},
    {"factor_set": "glec_accompanied", "scope": "WTW"},
    {"factor_set": "reference", "scope": "WTW"},
]


def test_scenarios_price_the_same_routes_without_rerouting(client, monkeypatch):
    """Routing costs seconds and seven OSRM calls; every extra pricing must be free."""
    calls = {"n": 0}
    original = route_module.road_route

    def counting(origin, destination):
        calls["n"] += 1
        return original(origin, destination)

    monkeypatch.setattr(route_module, "road_route", counting)
    payload = _post(client, scenarios=SCENARIOS).json()
    with_scenarios = calls["n"]

    calls["n"] = 0
    _post(client)
    assert with_scenarios == calls["n"]
    assert len(payload["scenarios"]) == len(SCENARIOS)


def test_each_scenario_reports_its_own_totals_and_sources(client):
    scenarios = {
        f"{s['factor_set']}|{s['scope']}": s for s in _post(client, scenarios=SCENARIOS).json()["scenarios"]
    }

    reference = scenarios["reference|TTW"]
    glec = scenarios["glec|TTW"]
    assert all("GLEC" in source for source in glec["sources"])
    assert not any("GLEC" in source for source in reference["sources"])

    # Same route, different basis: the ro-ro factor is what moves the answer.
    label = next(t["label"] for t in glec["totals"] if not t["is_all_road"])
    priced = {k: next(t["total_co2_kg"] for t in v["totals"] if t["label"] == label)
              for k, v in scenarios.items() if v["totals"]}
    assert priced["glec|TTW"] > priced["reference|TTW"] * 2
    assert priced["glec_accompanied|WTW"] > priced["glec|WTW"]


def test_an_unpriceable_scenario_carries_its_error_instead_of_vanishing(client):
    """The dashboard offers it as a choice, so it has to be able to say why it is empty."""
    scenarios = _post(client, scenarios=SCENARIOS).json()["scenarios"]
    broken = next(s for s in scenarios if s["factor_set"] == "reference" and s["scope"] == "WTW")

    assert broken["totals"] == []
    assert "no WTW factor" in broken["error"]
    assert broken["is_verified"] is False


def test_scenario_totals_omit_geometry_that_alternatives_already_carry(client):
    payload = _post(client, scenarios=SCENARIOS).json()
    totals = [t for s in payload["scenarios"] for t in s["totals"]]

    assert totals
    assert all("geometry" not in total for total in totals)


def test_a_request_without_scenarios_still_answers(client):
    assert _post(client).json()["scenarios"] == []


SHANGHAI = {"lon": 121.80, "lat": 31.23}
ROTTERDAM = {"lon": 4.13, "lat": 51.95}


def _compare(client, **overrides):
    body = {
        "origin": SHANGHAI,
        "destination": ROTTERDAM,
        "origin_country": "CN",
        "destination_country": "NL",
        "factor_set": "glec",
        "scope": "WTW",
    } | overrides
    return client.post("/api/compare", json=body)


def test_alternatives_report_whether_they_enter_a_listed_area(client):
    alternatives = _post(client).json()["alternatives"]

    for alternative in alternatives:
        risk = alternative["risk"]
        assert risk is not None
        # The pilot corridor's sea legs are tracked, so a clear result is a checked one.
        assert risk["untracked_sea_km"] == 0


def test_a_turkey_to_eu_sea_leg_is_billed_at_half_through_the_api(client):
    payload = _post(client, factor_set="glec", carbon_price_eur=80,
                    scenarios=[{"factor_set": "glec", "scope": "TTW"}]).json()
    totals = payload["scenarios"][0]["totals"]
    multimodal = next(t for t in totals if not t["is_all_road"])

    sea_leg = next(leg for leg in multimodal["ets"]["legs"])
    assert sea_leg["coverage_share"] == 0.5
    assert multimodal["ets"]["cost_eur"] > 0


def test_an_all_road_alternative_owes_no_allowances(client):
    payload = _post(client, scenarios=[{"factor_set": "glec", "scope": "TTW"}]).json()
    baseline = next(t for t in payload["scenarios"][0]["totals"] if t["is_all_road"])

    assert baseline["ets"]["cost_eur"] == 0
    assert baseline["ets"]["legs"] == []


def test_avoiding_suez_costs_distance_and_buys_out_of_the_listed_area(client):
    payload = _compare(client, avoid=["suez", "babalmandab", "panama"], surcharge_eur=4000).json()

    assert payload["direct"]["risk"]["is_exposed"]
    assert not payload["diverted"]["risk"]["is_exposed"]
    assert payload["extra_distance_km"] > 5000
    assert payload["extra_co2_kg"] > 0
    assert payload["avoided_zone_km"] > 1000


def test_the_surcharge_is_echoed_not_derived(client):
    """Premiums are negotiated against hull value; nothing here can compute one."""
    without = _compare(client).json()
    with_charge = _compare(client, surcharge_eur=4000).json()

    assert without["surcharge_eur"] == 0
    assert with_charge["total_extra_eur"] == pytest.approx(without["total_extra_eur"] + 4000)


def test_an_unblockable_passage_is_refused_with_the_list(client):
    response = _compare(client, avoid=["corinth"])

    assert response.status_code == 422
    assert "suez" in response.json()["detail"]


def test_a_diversion_that_cannot_be_sailed_says_so_instead_of_reporting_zero(client):
    """The Black Sea needs the Bosporus. Zero extras would read as a free reroute."""
    payload = _compare(
        client,
        destination={"lon": 28.65, "lat": 44.17},
        destination_country="RO",
        avoid=["gibraltar", "bosporus"],
    ).json()

    assert payload["direct"]["distance_km"] > 0
    assert "no sea route" in payload["diverted"]["unreachable"]
    assert payload["extra_distance_km"] is None
    assert payload["total_extra_eur"] is None


def _start_job(client, content=SHIPMENT_CSV, **fields):
    return client.post(
        "/api/report/jobs",
        files={"file": ("shipments.csv", content, "text/csv")},
        data={"scope": "TTW", "factor_set": "reference"} | fields,
    )


def _await_job(client, job_id, tries=200):
    for _ in range(tries):
        status = client.get(f"/api/report/jobs/{job_id}").json()
        if status["status"] in {"done", "failed"}:
            return status
        time.sleep(0.05)
    raise AssertionError(f"job {job_id} never finished")


def test_a_job_is_accepted_immediately_and_finishes_later(client):
    """A cold shipment costs seconds, so the upload cannot wait for the whole run."""
    response = _start_job(client)

    assert response.status_code == 202
    job = response.json()
    assert job["status"] in {"queued", "running"}
    assert job["total"] == 2

    finished = _await_job(client, job["id"])
    assert finished["status"] == "done"
    assert finished["done"] == finished["total"]
    assert finished["progress"] == 1.0


def test_the_finished_job_hands_back_the_same_report(client):
    job = _start_job(client, factor_set="glec", scope="WTW").json()
    _await_job(client, job["id"])

    response = client.get(f"/api/report/jobs/{job['id']}/file")
    assert response.status_code == 200
    assert "text/csv" in response.headers["content-type"]
    assert "factor set: glec" in response.text
    assert response.text.count("SEV-") == 2


def test_a_file_is_not_offered_before_the_job_finishes(client):
    """Handing back a half-written report would be worse than making the client wait."""
    job = _start_job(client).json()
    early = client.get(f"/api/report/jobs/{job['id']}/file")

    assert early.status_code in {200, 409}
    if early.status_code == 409:
        assert "not done" in early.json()["detail"]


def test_a_malformed_upload_fails_at_submission_not_a_minute_later(client):
    response = _start_job(client, content="origin_lon,origin_lat\n29.43,40.78\n")

    assert response.status_code == 422
    assert "missing column" in response.json()["detail"]


def test_a_job_that_cannot_price_anything_reports_failed(client):
    job = _start_job(client, factor_set="reference", scope="WTW").json()
    finished = _await_job(client, job["id"])

    assert finished["status"] == "failed"
    assert "no WTW factor" in finished["error"]
    assert client.get(f"/api/report/jobs/{job['id']}/file").status_code == 422


def test_an_unknown_job_is_a_404(client):
    assert client.get("/api/report/jobs/deadbeef").status_code == 404
    assert client.get("/api/report/jobs/deadbeef/file").status_code == 404


def test_a_place_search_returns_candidates_not_one_answer(client, monkeypatch):
    """Resolving a name behind the user's back is how a shipment lands in the wrong
    province; the endpoint hands back the choice instead.

    Nominatim is stubbed: this asserts our contract, not theirs, and a live call here
    made the suite depend on a rate-limited third party that CI cannot rely on.
    """
    from app.api import routes as routes_module
    from app.core.geocode import Candidate

    monkeypatch.setattr(
        routes_module,
        "search",
        lambda query, country=None, limit=5: [
            Candidate(name=f"Santa Maria {n}", lon=lon, lat=lat, kind="city", importance=i)
            for n, (lon, lat, i) in enumerate(
                [(-121.0, 34.9, 0.6), (-8.6, 41.2, 0.5), (13.4, 45.3, 0.4)]
            )
        ][:limit],
    )
    response = client.get("/api/places", params={"q": "Santa Maria", "limit": 5})

    assert response.status_code == 200
    places = response.json()
    assert len(places) > 1
    assert len({(p["lon"], p["lat"]) for p in places}) == len(places)
    for place in places:
        assert -180 <= place["lon"] <= 180 and -90 <= place["lat"] <= 90


def test_an_empty_place_query_is_refused(client):
    assert client.get("/api/places", params={"q": "   "}).status_code == 422


def test_reefer_is_absent_unless_the_cargo_is_refrigerated(client):
    payload = _post(client).json()

    assert all(a["reefer"] is None for a in payload["alternatives"])
    assert all(a["total_with_reefer_co2_kg"] is None for a in payload["alternatives"])


def test_reefer_is_additive_and_never_folded_into_the_transport_figure(client):
    """The transport number is published GLEC; the refrigeration number is derived.
    Merging them would hide which half rests on assumption."""
    dry = _post(client, factor_set="glec", scope="WTW").json()["alternatives"]
    cold = _post(client, factor_set="glec", scope="WTW", is_reefer=True).json()["alternatives"]

    for plain, chilled in zip(dry, cold):
        assert chilled["total_co2_kg"] == plain["total_co2_kg"]
        assert chilled["total_with_reefer_co2_kg"] > plain["total_co2_kg"]
        assert chilled["reefer"]["is_verified"] is False


def test_the_reefer_bill_grows_with_time_spent_not_distance_covered(client):
    """A multimodal route is slower, so it carries more refrigeration than the road
    baseline even where it covers similar ground. A per-kilometre model cannot see this."""
    alternatives = _post(client, is_reefer=True).json()["alternatives"]
    road = next(a for a in alternatives if a["is_all_road"])
    multimodal = next(a for a in alternatives if not a["is_all_road"])

    assert multimodal["timeline"]["total_hours"] > road["timeline"]["total_hours"]
    assert multimodal["reefer"]["co2_kg"] > road["reefer"]["co2_kg"]
    # The road leg never stops being driven, so nothing is billed as standing still.
    assert road["reefer"]["stationary_co2_kg"] == 0
    assert multimodal["reefer"]["stationary_co2_kg"] > 0


def test_a_derived_reefer_figure_carries_its_warning_into_every_scenario(client):
    payload = _post(
        client, is_reefer=True, scenarios=[{"factor_set": "glec", "scope": "WTW"}]
    ).json()
    total = payload["scenarios"][0]["totals"][0]

    assert total["reefer"] is not None
    assert any("derived, not published" in w for w in total["reefer"]["warnings"])


def _stub_catchment(monkeypatch, cells=None, notes=None):
    from app.api import routes as routes_module
    from app.core.catchment import Catchment, CatchmentCell

    cells = cells if cells is not None else [
        CatchmentCell(lon=29.0, lat=40.0, terminal_id="pendik", duration_h=1.2),
        CatchmentCell(lon=30.0, lat=40.0, terminal_id="yalova", duration_h=2.5),
    ]
    monkeypatch.setattr(
        routes_module, "build_catchment",
        lambda **kwargs: Catchment(
            cells=cells, spacing_deg=kwargs.get("spacing_deg", 1.0),
            bounds=(26.0, 36.0, 45.0, 42.0),
            max_duration_h=kwargs.get("max_duration_h", 8.0),
            sampled=10, unreachable=8, notes=notes or ["ornekleme notu"],
        ),
    )


def test_the_catchment_says_how_coarsely_it_was_sampled(client, monkeypatch):
    """The spacing is part of the answer: a client that drew a smooth boundary over
    these cells would claim a precision that was never measured."""
    _stub_catchment(monkeypatch)

    body = client.get("/api/catchment", params={"spacing_deg": 1.0}).json()

    assert body["spacing_deg"] == 1.0
    assert body["sampled"] == 10 and body["unreachable"] == 8
    assert body["cells_by_terminal"] == {"pendik": 1, "yalova": 1}
    assert body["notes"]


def test_a_spacing_finer_than_the_engine_allows_is_refused(client):
    assert client.get("/api/catchment", params={"spacing_deg": 0.01}).status_code == 422


def test_a_nonsense_time_limit_is_refused(client):
    assert client.get("/api/catchment", params={"max_duration_h": 0}).status_code == 422


def test_an_unroutable_catchment_reports_rather_than_crashing(client, monkeypatch):
    from app.api import routes as routes_module
    from app.core.road import RoadRoutingError

    def unroutable(**kwargs):
        raise RoadRoutingError("table limit reached")

    monkeypatch.setattr(routes_module, "build_catchment", unroutable)

    response = client.get("/api/catchment")
    assert response.status_code == 422
    assert "table limit" in response.json()["detail"]


@pytest.mark.parametrize(
    "output_format,magic,media",
    [
        ("csv", b"# FreightPrint", "text/csv"),
        ("xlsx", b"PK", "spreadsheetml"),
        ("pdf", b"%PDF-", "application/pdf"),
    ],
)
def test_the_report_is_served_in_the_format_that_was_asked_for(
    client, output_format, magic, media
):
    """A spreadsheet handed back labelled as text opens as gibberish, so the media type
    has to follow the format and not the other way round."""
    response = _upload(client, output_format=output_format)

    assert response.status_code == 200
    assert response.content.startswith(magic)
    assert media in response.headers["content-type"]
    assert f".{output_format}" in response.headers["content-disposition"]


def test_an_unknown_report_format_is_refused_rather_than_defaulted(client):
    response = _upload(client, output_format="docx")

    assert response.status_code == 422
    assert "csv" in response.json()["detail"]


def test_every_format_states_the_same_basis(client):
    """The point of the deliverable formats is that the basis travels with them."""
    from io import BytesIO

    from openpyxl import load_workbook
    from pypdf import PdfReader

    csv = _upload(client, factor_set="glec", scope="WTW", output_format="csv").text
    xlsx = _upload(client, factor_set="glec", scope="WTW", output_format="xlsx").content
    pdf = _upload(client, factor_set="glec", scope="WTW", output_format="pdf").content

    assert "glec" in csv and "GLEC Framework" in csv

    workbook = load_workbook(BytesIO(xlsx))
    basis = " ".join(
        str(c.value) for row in workbook["Esas ve kaynaklar"].iter_rows() for c in row if c.value
    )
    assert "glec" in basis and "GLEC Framework" in basis

    text = " ".join((page.extract_text() or "") for page in PdfReader(BytesIO(pdf)).pages)
    assert "glec" in text and "GLEC Framework" in text


def test_a_background_job_returns_the_format_it_was_started_with(client):
    started = client.post(
        "/api/report/jobs",
        files={"file": ("shipments.csv", SHIPMENT_CSV, "text/csv")},
        data={"scope": "TTW", "factor_set": "reference", "output_format": "xlsx"},
    )
    assert started.status_code == 202
    job = started.json()
    assert job["filename"].endswith(".xlsx")

    for _ in range(100):
        status = client.get(f"/api/report/jobs/{job['id']}").json()
        if status["status"] in {"done", "failed"}:
            break
        time.sleep(0.1)

    assert status["status"] == "done", status.get("error")
    downloaded = client.get(f"/api/report/jobs/{job['id']}/file")
    assert downloaded.content.startswith(b"PK")
    assert "spreadsheetml" in downloaded.headers["content-type"]


def test_the_api_says_which_road_fuels_exist(client):
    """The names are not guessable — the rows are diesel_b5 and electric_tr, so a caller
    reaching for "diesel" or "electric" gets an error. A factor row nobody can discover
    is a factor row nobody can use."""
    sets = {s["name"]: s for s in client.get("/api/factor-sets").json()}
    fuels = {f["fuel_type"]: f for f in sets["glec"]["road_fuels"]}

    assert {"diesel_b5", "hvo_uco", "hvo_palm", "electric_tr"} <= set(fuels)
    assert fuels["diesel_b5"]["is_default"] and fuels["diesel_b5"]["is_verified"]
    assert not fuels["hvo_uco"]["is_verified"], "a derived row must not look published"
    assert fuels["hvo_uco"]["label"] != "hvo_uco", "no human-readable label"


def test_the_listed_fuels_are_exactly_the_ones_that_price(client):
    """The list is generated from the factor file, so it cannot drift from what the
    engine will actually accept."""
    sets = {s["name"]: s for s in client.get("/api/factor-sets").json()}

    for fuel in sets["glec"]["road_fuels"]:
        response = _post(
            client, factor_set="glec", scope="WTW", road_fuel_type=fuel["fuel_type"]
        )
        assert response.status_code == 200, f"{fuel['fuel_type']} is listed but will not price"


def test_the_feedstock_choice_reaches_the_answer(client):
    """Selecting palm-oil HVO instead of waste cooking oil has to move the number, or
    the selector is decoration."""
    uco = _post(client, factor_set="glec", scope="WTW", road_fuel_type="hvo_uco").json()
    palm = _post(client, factor_set="glec", scope="WTW", road_fuel_type="hvo_palm").json()

    road_uco = next(a for a in uco["alternatives"] if a["is_all_road"])["total_co2_kg"]
    road_palm = next(a for a in palm["alternatives"] if a["is_all_road"])["total_co2_kg"]

    assert road_palm > road_uco * 3


def test_the_portfolio_groups_shipments_into_lanes(client):
    response = client.post(
        "/api/portfolio",
        files={"file": ("shipments.csv", SHIPMENT_CSV, "text/csv")},
        data={"scope": "WTW", "factor_set": "glec"},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["lanes"]
    assert body["total_co2_kg"] > 0
    assert len(body["tested_sets"]) > 1, "robustness against one basis is not robustness"


def test_every_lane_says_which_bases_it_wins_under(client):
    """The column that decides whether a saving can be defended, so it must never be
    absent or silently empty when the lane claims to be robust."""
    body = client.post(
        "/api/portfolio",
        files={"file": ("shipments.csv", SHIPMENT_CSV, "text/csv")},
        data={"scope": "WTW", "factor_set": "glec"},
    ).json()

    for lane in body["lanes"]:
        assert lane["tested_under"]
        assert set(lane["wins_under"]) <= set(lane["tested_under"])
        if lane["is_robust"]:
            assert set(lane["wins_under"]) == set(lane["tested_under"])
        assert not (lane["is_robust"] and lane["is_contested"])


def test_a_portfolio_from_a_bad_file_is_refused(client):
    response = client.post(
        "/api/portfolio",
        files={"file": ("shipments.csv", "origin_lon,origin_lat\n29.43,40.78\n", "text/csv")},
        data={"scope": "WTW", "factor_set": "glec"},
    )

    assert response.status_code == 422
    assert "missing column" in response.json()["detail"]


def test_the_conformance_endpoint_reports_the_gaps(client):
    body = client.get("/api/conformance", params={"factor_set": "glec", "scope": "WTW"}).json()

    assert body["verdict"] in {"reportable", "reportable-not-verifiable", "not-reportable"}
    ids = {c["id"] for c in body["checks"]}
    assert {"hub_emissions", "primary_data"} <= ids
    for check in body["checks"]:
        if check["id"] in {"hub_emissions", "primary_data"}:
            assert check["status"] == "missing", "a gap must never quietly start passing"


def test_a_tank_to_wheel_basis_is_reported_unreportable(client):
    body = client.get("/api/conformance", params={"factor_set": "glec", "scope": "TTW"}).json()

    assert body["verdict"] == "not-reportable"
    assert any(c["is_blocking"] for c in body["checks"])


def test_a_basis_that_cannot_price_is_refused(client):
    response = client.get("/api/conformance", params={"factor_set": "reference", "scope": "WTW"})

    assert response.status_code == 422


def test_the_co2_toll_component_reaches_the_dashboard(client):
    body = _post(client, factor_set="glec", scope="WTW",
                 scenarios=[{"factor_set": "glec", "scope": "WTW"}]).json()
    totals = body["scenarios"][0]["totals"]

    road = next(t for t in totals if t["is_all_road"])
    assert road["co2_toll"]["total_eur"] > 0
    assert any(c["iso"] == "DE" and c["priced"] for c in road["co2_toll"]["countries"])


def test_a_country_without_a_published_carbon_price_is_unpriced_not_free(client):
    body = _post(client, factor_set="glec", scope="WTW",
                 scenarios=[{"factor_set": "glec", "scope": "WTW"}]).json()
    road = next(t for t in body["scenarios"][0]["totals"] if t["is_all_road"])

    austria = next(c for c in road["co2_toll"]["countries"] if c["iso"] == "AT")
    assert not austria["priced"]
    assert austria["cost_eur"] == 0
    assert austria["reason"], "an unpriced country must say why"


def test_staying_off_german_roads_shows_up_as_money(client):
    body = _post(client, factor_set="glec", scope="WTW",
                 scenarios=[{"factor_set": "glec", "scope": "WTW"}]).json()
    totals = body["scenarios"][0]["totals"]

    road = next(t for t in totals if t["is_all_road"])
    multimodal = next(t for t in totals if not t["is_all_road"])

    assert road["co2_toll"]["total_eur"] > multimodal["co2_toll"]["total_eur"] * 5


# ── the sea factor reaches the response, not only its own tests ───────────────

def _multimodal(client, **overrides):
    """The alternative that actually sails; an all-road option has no sea factor."""
    payload = _post(client, **overrides).json()
    return next(a for a in payload["alternatives"] if not a["is_all_road"])


def test_the_sea_factor_comparison_reaches_the_api(client):
    """The benchmark is only worth building if a caller can see it. This module was once
    complete, correct, and reachable from nothing but its own test file."""
    sea = _multimodal(client)["sea_factor"]

    assert sea is not None
    assert sea["ships"] > 200
    assert "THETIS-MRV" in sea["observed_source"]


def test_an_all_road_option_carries_no_sea_factor(client):
    """Nothing sails, so there is nothing to check. A comparison offered here would be
    about a leg the route does not have."""
    payload = _post(client).json()
    road = next(a for a in payload["alternatives"] if a["is_all_road"])

    assert road["sea_factor"] is None


def test_the_comparison_uses_the_ttw_row_even_when_the_shipment_is_priced_wtw(client):
    """MRV reports the CO2 a ship emitted from fuel it burned. Comparing GLEC's
    well-to-wake 0.068 against it charges the observation for fuel production nobody
    measured, and makes the factor look about 8% worse than it is."""
    sea = _multimodal(client, scope="WTW", factor_set="glec")["sea_factor"]

    assert sea["compared_scope"] == "TTW"
    assert sea["factor"] == pytest.approx(0.063)
    assert any("TTW" in note and "WTW" in note for note in sea["notes"]), \
        "priced on one basis and compared on another without saying so"


def test_the_factor_that_changes_changes_the_comparison(client):
    """A panel that reports the same verdict whatever the engine is told to use is
    decoration. The accompanied basis is half again the trailer one."""
    trailer = _multimodal(client, factor_set="glec")["sea_factor"]
    accompanied = _multimodal(client, factor_set="glec_accompanied")["sea_factor"]

    assert accompanied["factor"] > trailer["factor"]
    assert accompanied["ratio"] > trailer["ratio"]
    assert trailer["compared_row"] != accompanied["compared_row"]


def test_a_factor_describing_traffic_the_fleet_does_not_carry_says_so(client):
    """Accompanied haulage largely sails ro-pax, and no ro-pax ship reports mass-based
    transport work — the whole class is absent from the observation. Reporting the
    comparison anyway is fine; reporting it as a test would not be."""
    accompanied = _multimodal(client, factor_set="glec_accompanied")["sea_factor"]

    assert accompanied["is_comparable"] is False
    assert any("ro-pax" in note for note in accompanied["notes"])

    trailer = _multimodal(client, factor_set="glec")["sea_factor"]
    assert trailer["is_comparable"] is True


def test_the_observed_fleet_is_named_ship_by_ship_type(client):
    """So a reader can see that ro-pax is missing rather than assume it was counted."""
    sea = _multimodal(client)["sea_factor"]

    assert "Ro-ro ship" in sea["ship_types"]
    assert "Ro-pax ship" not in sea["ship_types"]
    assert sum(sea["ship_types"].values()) == sea["ships"]


def test_the_spread_travels_with_the_verdict(client):
    """The pass is the smaller finding. The middle half of the fleet spans a factor of
    about 2.7, which says no fleet average describes the ship carrying the load."""
    sea = _multimodal(client, factor_set="glec")["sea_factor"]

    assert sea["verdict"] == "within"
    assert sea["q1"] < sea["factor"] < sea["q3"]
    assert sea["spread"] > 2


def test_the_validation_datasets_own_sea_factor_is_below_every_verified_ship(client):
    """The sharpest thing this benchmark found, and it is about the default basis.

    The reference set is not a published standard — it is the figure the validation
    dataset's own carbon report used, kept here so the engine can be checked against a
    real document. Its sea value is 0.012, and the emission factor table has always
    carried a note that this looked closer to a container ship than to a ro-ro service.

    That was a suspicion with nothing behind it. Against EU MRV it is now an observation:
    not one of the verified ro-ro ships in the reporting period is that clean, so a
    report priced on this basis understates its sea leg by roughly a factor of four.
    """
    sea = _multimodal(client)["sea_factor"]

    assert sea["compared_row"].startswith("reference"), "the default basis changed"
    assert sea["verdict"] == "below"
    assert sea["share_below"] == 0.0, "some ship is now that clean; rewrite this finding"
    assert sea["ratio"] < 0.3


# ── where a distance came from, not just where the factor did ─────────────────

def test_a_leg_says_where_its_distance_came_from(client):
    """A sea leg priced from the carrier's own figure and one checked against a published
    table are different claims, and the kilometres look identical either way.

    NGA Pub. 151 puts every sea leg on this corridor 9-26% above the carrier's number, so
    "nobody has checked this" is not a formality here.
    """
    sea = next(leg for leg in _multimodal(client)["legs"] if leg["mode"] == "sea")

    assert sea["distance_source"], "the sea distance arrives with no provenance"
    assert sea["distance_is_verified"] is False, (
        "a sea distance is marked verified while the service table still says otherwise"
    )


def test_a_computed_road_distance_claims_no_service_table_source(client):
    """Road distance is routed here by OSRM rather than read from `service_legs.csv`, so
    borrowing that file's provenance would credit it to a source it never came from."""
    payload = _post(client).json()
    road = next(a for a in payload["alternatives"] if a["is_all_road"])["legs"][0]

    assert road["distance_source"] == ""


# ── the sea distance beside the publication that surveys it ───────────────────

def test_the_sea_distance_comparison_reaches_a_caller(client):
    """Sea is 87% of this corridor's emissions and its distance had one cross-check in
    total. Six now exist and every one says the service table reads high."""
    legs = _multimodal(client)["sea_distances"]

    assert legs, "the comparison built for the biggest lever reaches nobody"
    leg = legs[0]
    assert leg["published_km"] < leg["engine_km"]
    assert leg["delta_pct"] > 9


def test_the_engine_still_prices_with_the_carriers_figure(client):
    """Chosen deliberately, and the same footing as every other external check here: the
    observation is reported beside the assumption, never substituted for it. An engine
    that quietly swapped in a survey taken for a different purpose would stop reproducing
    the report it is validated against."""
    alternative = _multimodal(client)
    sea_leg = next(leg for leg in alternative["legs"] if leg["mode"] == "sea")
    comparison = alternative["sea_distances"][0]

    assert sea_leg["distance_km"] == comparison["engine_km"]
    assert sea_leg["distance_km"] != comparison["published_km"]


def test_the_published_half_of_the_comparison_is_declared(client):
    """Pub 151 lists neither Pendik nor Yalova, so Istanbul stands for them and the hop is
    measured from coordinates. Small here — but a figure that stops saying so stops being
    one, which is how the Pendik offset was assumed at twice its real size."""
    leg = _multimodal(client)["sea_distances"][0]

    assert leg["published_nm"] > 0
    assert leg["estimated_share"] < 0.10
    assert leg["is_representative"] is True
    assert "Trieste" in leg["source_line"], "the line it was read from is not carried"


def test_an_all_road_option_has_no_sea_distance_to_compare(client):
    payload = _post(client).json()
    road = next(a for a in payload["alternatives"] if a["is_all_road"])

    assert road["sea_distances"] == []
