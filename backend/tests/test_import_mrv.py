"""The THETIS-MRV importer, which is the one piece of this project a person has to run.

EMSA publishes verified per-ship CO2 for every vessel over 5,000 GT calling at an EEA
port — the only independent observation of the ro-ro figure this project's headline
rests on. It cannot be fetched: the portal is a JavaScript application, the export sits
behind a reCAPTCHA, and no direct file URL is published.

So the importer takes a file somebody downloaded, and what these tests defend is that it
never guesses. The stand-in workbook below is shaped like the real publication — two
rows of grouping labels above the header, and columns that differ only by the words "on
laden voyages" — because both of those broke the first version.
"""

import csv
import importlib.util
import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parents[2]
spec = importlib.util.spec_from_file_location("import_mrv", REPO / "scripts" / "import_mrv.py")
import_mrv = importlib.util.module_from_spec(spec)
sys.modules["import_mrv"] = import_mrv
spec.loader.exec_module(import_mrv)

# Row 1 and 2 are grouping labels; the real header is row 3. Reading row 1 as the header
# finds nothing at all, which is what the guard reported against the first real export.
GROUP_ROW_1 = ["Ship", None, None, None, "Annual monitoring results"]
GROUP_ROW_2 = [None, None, None, None, "Average energy efficiency"]
HEADER = [
    "IMO Number", "Name", "Ship type", "Reporting Period",
    "CO₂ emissions per transport work (mass) [g CO₂ / m tonnes · n miles]",
    "CO₂ emissions per transport work (mass) on laden voyages [g CO₂ / m tonnes · n miles]",
    "CO₂eq emissions per transport work (mass) [g CO₂eq / m tonnes · n miles]",
    "Fuel consumption per transport work (mass) [g / m tonnes · n miles]",
]
SHIPS = [
    (9000001, "A", "Ro-ro ship", 2024, 25.4, 20.1, 26.9, 8.0),
    (9000002, "B", "Ro-pax ship", 2024, 41.9, 33.0, 44.4, 13.2),
    (9000003, "C", "Container ship", 2024, 8.1, 6.4, 8.6, 2.6),
    (9000004, "D", "Ro-ro ship", 2024, 31.2, 25.0, 33.1, 9.8),
    # MRV writes this literal string where a ship reported no transport work.
    (9000005, "E", "Ro-ro ship", 2024, "Division by zero!", None, None, None),
    (9000006, "F", "Vehicle carrier", 2024, 12.0, 9.5, 12.7, 3.8),
    (9000007, "G", "Bulk carrier", 2024, 3.4, 2.7, 3.6, 1.1),
]


def _workbook(tmp_path, header=HEADER, ships=SHIPS, sheet="2024 Full ERs"):
    book = openpyxl.Workbook()
    page = book.active
    page.title = sheet
    page.append(GROUP_ROW_1)
    page.append(GROUP_ROW_2)
    page.append(header)
    for ship in ships:
        page.append(ship)
    path = tmp_path / "mrv.xlsx"
    book.save(path)
    return path


@pytest.fixture
def workbook(tmp_path, monkeypatch):
    monkeypatch.setattr(import_mrv, "OUT", tmp_path / "roro.csv")
    return _workbook(tmp_path)


def rows_written():
    return list(csv.DictReader(import_mrv.OUT.open(encoding="utf-8")))


def test_the_header_is_found_rather_than_assumed_to_be_the_first_row(tmp_path):
    """Two rows of grouping labels sit above it in the real publication."""
    book = openpyxl.load_workbook(_workbook(tmp_path), read_only=True)
    number, header = import_mrv.header_row(book["2024 Full ERs"])
    book.close()

    assert number == 3
    assert header[0] == "IMO Number"


def test_a_workbook_with_no_header_says_so_instead_of_reading_row_one(tmp_path):
    book = openpyxl.Workbook()
    book.active.append(["something", "else"])
    path = tmp_path / "odd.xlsx"
    book.save(path)

    opened = openpyxl.load_workbook(path, read_only=True)
    with pytest.raises(LookupError, match="imo number"):
        import_mrv.header_row(opened["Sheet"])
    opened.close()


def test_the_laden_only_column_is_not_mistaken_for_the_full_one():
    """They differ by four words and are different numbers. Matching on a substring
    alone picks whichever comes first, which is why the rules carry exclusions."""
    found, missing = import_mrv.find_columns(HEADER)

    assert found["co2_per_tw"] == 4
    assert found["co2_per_tw_laden"] == 5
    assert "co2_per_tw_freight" in missing, "a column that is absent must be reported"


def test_the_co2eq_and_fuel_columns_are_kept_out():
    """Both contain the same phrase. CO2-equivalent includes methane and nitrous oxide,
    and fuel consumption is not an emission at all."""
    found, _ = import_mrv.find_columns(HEADER)

    assert found["co2_per_tw"] not in (6, 7)


def test_a_missing_required_column_stops_the_run(tmp_path, monkeypatch, capsys):
    """Guessing at a header is exactly the shortcut this project does not take."""
    monkeypatch.setattr(import_mrv, "OUT", tmp_path / "out.csv")
    thin = _workbook(tmp_path, header=["IMO Number", "Name"], ships=[(1, "A")])

    assert import_mrv.derive(thin) == 0
    assert "bulunamadi" in capsys.readouterr().err
    assert not (tmp_path / "out.csv").exists(), "wrote a file it could not fill honestly"


def test_only_ro_ro_vessels_are_kept(workbook):
    """Table 45 describes a ro-ro fleet. A vehicle carrier moves cars and a bulk carrier
    moves ore; either would be comparing different ships and calling it validation."""
    assert import_mrv.derive(workbook) == 3

    kept = {row["ship_type"] for row in rows_written()}
    assert kept == {"Ro-ro ship", "Ro-pax ship"}


def test_a_ship_that_reported_no_transport_work_is_dropped_not_zeroed(workbook):
    """Read as a number "Division by zero!" is nothing, and nothing would drag the
    median to the floor."""
    import_mrv.derive(workbook)

    assert "9000005" not in {row["imo"] for row in rows_written()}


def test_the_unit_conversion_is_correct(workbook):
    """MRV reports grams per tonne-nautical-mile, GLEC kilograms per tonne-kilometre.
    Comparing them unconverted is wrong by a factor of 1,852."""
    import_mrv.derive(workbook)
    row = next(r for r in rows_written() if r["imo"] == "9000001")

    assert float(row["kg_co2_per_tonne_km"]) == pytest.approx(25.4 / 1.852 / 1000, rel=1e-4)


def test_re_importing_a_period_replaces_it_rather_than_doubling_the_fleet(workbook):
    """A corrected publication is re-imported over the old one; running it twice must
    not report twice as many ships."""
    import_mrv.derive(workbook)
    import_mrv.derive(workbook)

    assert len(rows_written()) == 3


def test_importing_another_period_adds_to_it(workbook, tmp_path):
    """Each year is imported in turn and the file accumulates."""
    import_mrv.derive(workbook)
    older = _workbook(
        tmp_path,
        ships=[(9000001, "A", "Ro-ro ship", 2023, 28.0, 22.0, 29.6, 8.8)],
        sheet="2023 Full ERs",
    )
    import_mrv.derive(older)

    periods = {row["reporting_period"] for row in rows_written()}
    assert periods == {"2023", "2024"}


def test_the_full_reports_sheet_is_preferred_over_the_partial_one(tmp_path, monkeypatch):
    """A partial return covers part of a year, so its intensity is not comparable with
    an annual figure."""
    monkeypatch.setattr(import_mrv, "OUT", tmp_path / "out.csv")
    book = openpyxl.Workbook()
    partial = book.active
    partial.title = "2024 Partial ERs"
    for row in (GROUP_ROW_1, GROUP_ROW_2, HEADER):
        partial.append(row)
    partial.append((9000099, "P", "Ro-ro ship", 2024, 99.0, 88.0, 105.0, 31.0))
    full = book.create_sheet("2024 Full ERs")
    for row in (GROUP_ROW_1, GROUP_ROW_2, HEADER):
        full.append(row)
    full.append((9000001, "A", "Ro-ro ship", 2024, 25.4, 20.1, 26.9, 8.0))
    path = tmp_path / "both.xlsx"
    book.save(path)

    import_mrv.derive(path)

    assert {row["imo"] for row in rows_written()} == {"9000001"}


def test_running_it_with_no_file_explains_where_to_get_one(capsys, monkeypatch):
    """The one step a person has to take, so the instruction lives in the tool."""
    monkeypatch.setattr(sys, "argv", ["import_mrv.py"])

    assert import_mrv.main() == 1
    out = capsys.readouterr().out
    assert "mrv.emsa.europa.eu" in out
    assert "reCAPTCHA" in out, "does not say why it cannot fetch the file itself"
