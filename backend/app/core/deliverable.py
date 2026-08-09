"""The report in the forms a customer will actually accept: Excel and PDF.

CSV is the interchange format; neither of these replaces it. Excel is what a logistics
department edits and files, PDF is what gets attached to a customer's own reporting.

Whatever the format, the same rule holds as in the CSV writer: **a carbon figure without
the basis it was produced on cannot be checked by whoever receives it.** The factor set,
the scope, the sources and every warning travel with the numbers in both formats, and a
shipment that could not be routed keeps its reason instead of printing as a blank row.
"""

import io
import os
from datetime import date

import reportlab
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from .report import OUTPUT_COLUMNS, Report
from .uncertainty import round_to_significant

# Bitstream Vera ships inside reportlab and covers the Turkish alphabet in full. The
# built-in Helvetica does not: reportlab's standard fonts are Latin-1, which has no
# dotless i, no s-cedilla, no g-breve and no dotted capital I, so a Turkish report set
# in them comes out broken. Using the bundled file keeps the container free of any
# system font dependency.
_FONT_DIR = os.path.join(os.path.dirname(reportlab.__file__), "fonts")
FONT_REGULAR, FONT_BOLD = "FP-Sans", "FP-Sans-Bold"

TITLE = "FreightPrint — Karbon Raporu"

HEADER_TR = {
    "reference": "Referans",
    "origin_name": "Kalkış",
    "destination_name": "Varış",
    "tonnage": "Ton",
    "route": "Rota",
    "road_km": "Karayolu km",
    "sea_km": "Deniz km",
    "rail_km": "Demiryolu km",
    "total_km": "Toplam km",
    "road_co2_kg": "Karayolu kg",
    "sea_co2_kg": "Deniz kg",
    "rail_co2_kg": "Demiryolu kg",
    "total_co2_kg": "Toplam kg CO2",
    "all_road_co2_kg": "Tam karayolu kg",
    "saving_co2_kg": "Fark kg",
    "trees_equivalent": "Ağaç eşdeğeri",
    "status": "Durum",
}

# Columns the PDF carries. The per-mode breakdown stays in the spreadsheet: a summary
# that needs a magnifying glass is not a summary. `status` is not optional though — a
# shipment that could not be routed otherwise prints as a row of blanks, and a reader
# has no way to tell that from a shipment with no emissions.
PDF_COLUMNS = (
    "reference", "origin_name", "destination_name", "tonnage",
    "route", "total_km", "total_co2_kg", "all_road_co2_kg", "saving_co2_kg", "status",
)

STATUS_TR = {"ok": "hesaplandı"}


def _number_tr(value: float, decimals: int = 0) -> str:
    """A number written the way the rest of this document is: in Turkish.

    This is not cosmetic. Turkish groups thousands with a full stop and marks decimals
    with a comma, so the English default renders 4770 kg as "4,770" — which a Turkish
    reader parses as four point seven seven. In a carbon report that is a factor of a
    thousand, in the direction of looking harmless.
    """
    formatted = f"{value:,.{decimals}f}"
    # Swap the two separators via a placeholder so neither pass undoes the other.
    return formatted.replace(",", "\x00").replace(".", ",").replace("\x00", ".")


def _cell_text_for(column: str, value) -> str:
    """One PDF cell's text, formatted for what the column holds."""
    if value is None:
        return ""
    if column == "status":
        return STATUS_TR.get(value, value.replace("failed:", "hesaplanamadı:"))
    if column == "tonnage":
        # Excel hands every number back as a float, so 24 tonnes arrives as 24.0.
        return _number_tr(value, 0 if float(value).is_integer() else 1)
    if column.endswith("_km"):
        return _number_tr(value, 0)
    if column.endswith("_kg") or column == "trees_equivalent":
        return _number_tr(value, 0)
    return str(value)


_fonts_registered = False


def _register_fonts() -> None:
    global _fonts_registered
    if _fonts_registered:
        return
    pdfmetrics.registerFont(TTFont(FONT_REGULAR, os.path.join(_FONT_DIR, "Vera.ttf")))
    pdfmetrics.registerFont(TTFont(FONT_BOLD, os.path.join(_FONT_DIR, "VeraBd.ttf")))
    pdfmetrics.registerFontFamily(FONT_REGULAR, normal=FONT_REGULAR, bold=FONT_BOLD)
    _fonts_registered = True


def _row_values(report_row) -> dict:
    """One report row as the values both formats print, keyed by column name."""
    row = report_row
    return {
        "reference": row.shipment.reference,
        "origin_name": row.shipment.origin_name,
        "destination_name": row.shipment.destination_name,
        "tonnage": row.shipment.tonnage,
        "route": row.route_label,
        "road_km": round(row.distance_by_mode.get("road", 0.0), 1),
        "sea_km": round(row.distance_by_mode.get("sea", 0.0), 1),
        "rail_km": round(row.distance_by_mode.get("rail", 0.0), 1),
        "total_km": round(row.total_km, 1),
        "road_co2_kg": round_to_significant(row.co2_by_mode.get("road", 0.0)),
        "sea_co2_kg": round_to_significant(row.co2_by_mode.get("sea", 0.0)),
        "rail_co2_kg": round_to_significant(row.co2_by_mode.get("rail", 0.0)),
        "total_co2_kg": round_to_significant(row.total_co2_kg),
        "all_road_co2_kg": (
            round_to_significant(row.all_road_co2_kg) if row.all_road_co2_kg else None
        ),
        "saving_co2_kg": (
            round_to_significant(row.saving_co2_kg) if row.saving_co2_kg is not None else None
        ),
        "trees_equivalent": round(row.trees) if row.trees is not None else None,
        "status": row.status,
    }


def _basis_lines(report: Report) -> list[tuple[str, str]]:
    """The provenance block, identical in both formats."""
    return [
        ("Faktör seti", report.factor_set),
        ("Kapsam", f"{report.scope} ({'kuyudan tekere' if report.scope == 'WTW' else 'depodan tekere'})"),
        ("Kaynaklar", "; ".join(report.sources) or "DOĞRULANMAMIŞ"),
        ("Üretim tarihi", date.today().isoformat()),
        ("Hesaplanan sevkiyat", str(len(report.calculated))),
        ("Hesaplanamayan", str(len(report.failed))),
    ]


METHOD_NOTE = (
    "Bu rapor, sevkiyatları kalkış ve varış noktalarından gerçek karayolu/deniz/demiryolu "
    "ağları üzerinden rotalayarak üretilmiştir. Her sevkiyat için en düşük emisyonlu "
    "seçenek raporlanır — bu tam karayolu da olabilir. Emisyon = mesafe × ton × faktör; "
    "faktör, yukarıda adı geçen sette yayımlanmış değerdir ve o kaynağın kendi doluluk "
    "ve boş dönüş varsayımlarını zaten içerir. \"Fark\" sütunu, seçilen rotanın tam "
    "karayolu alternatifine göre kazancıdır. Çok modlu seçeneğin bu esas altında daha "
    "fazla emisyon ürettiği durumlarda en düşük emisyonlu seçenek zaten tam karayolu "
    "olur; o satırlarda rota \"all-road\" görünür ve fark sıfırdır. Yani sıfır fark, "
    "hesabın yapılmadığı değil, çok modlu alternatifin bu esasta kazandırmadığı "
    "anlamına gelir."
)


def report_to_xlsx(report: Report) -> bytes:
    """The report as a workbook: data on one sheet, its basis on another.

    The basis is a sheet rather than a footnote because a spreadsheet gets sorted,
    filtered and pasted from, and anything sitting above the header row is the first
    thing to be lost when it does.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Rapor"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1B5FB0")
    sheet.append([HEADER_TR[column] for column in OUTPUT_COLUMNS])
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # Freeze the header: these files are read by scrolling, and a number whose column
    # has scrolled out of sight is a number without units.
    sheet.freeze_panes = "A2"

    # No special styling for a negative difference: the report always carries the
    # lowest-emission option, so the gap against the all-road baseline cannot be
    # negative here. Where multimodal loses, that baseline simply *is* the chosen
    # route and the difference is zero — visible in the route column, not hidden.
    for report_row in report.rows:
        values = _row_values(report_row)
        sheet.append([values[column] for column in OUTPUT_COLUMNS])

    for index, column in enumerate(OUTPUT_COLUMNS, start=1):
        letter = get_column_letter(index)
        longest = max(
            [len(HEADER_TR[column])]
            + [len(str(_row_values(r)[column] or "")) for r in report.rows]
        )
        sheet.column_dimensions[letter].width = min(max(longest + 2, 10), 34)
        if column.endswith(("_km", "_kg", "tonnage", "equivalent")):
            for cell in sheet[letter][1:]:
                cell.number_format = "#,##0.0" if column.endswith("_km") else "#,##0"

    sheet.auto_filter.ref = f"A1:{get_column_letter(len(OUTPUT_COLUMNS))}{sheet.max_row}"

    basis = workbook.create_sheet("Esas ve kaynaklar")
    basis.column_dimensions["A"].width = 24
    basis.column_dimensions["B"].width = 96
    basis.append([TITLE])
    basis["A1"].font = Font(bold=True, size=14)
    basis.append([])
    for label, value in _basis_lines(report):
        basis.append([label, value])
        basis.cell(row=basis.max_row, column=1).font = Font(bold=True)
    basis.append([])
    basis.append(["Toplam CO2 (kg)", round_to_significant(report.total_co2_kg)])
    basis.append(["Toplam fark (kg)", round_to_significant(report.total_saving_co2_kg)])
    basis.append([])
    if report.warnings:
        basis.append(["Uyarılar", ""])
        basis.cell(row=basis.max_row, column=1).font = Font(bold=True, color="B3261E")
        for warning in report.warnings:
            basis.append(["", warning])
            basis.cell(row=basis.max_row, column=2).alignment = Alignment(wrap_text=True)
        basis.append([])
    basis.append(["Yöntem", METHOD_NOTE])
    basis.cell(row=basis.max_row, column=1).font = Font(bold=True)
    basis.cell(row=basis.max_row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    basis.row_dimensions[basis.max_row].height = 90

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _pdf_styles() -> dict:
    _register_fonts()
    sample = getSampleStyleSheet()
    return {
        "title": ParagraphStyle(
            "fp-title", parent=sample["Title"], fontName=FONT_BOLD, fontSize=17, spaceAfter=2,
        ),
        "body": ParagraphStyle(
            "fp-body", parent=sample["BodyText"], fontName=FONT_REGULAR,
            fontSize=8.2, leading=11.5, alignment=TA_LEFT,
        ),
        "warn": ParagraphStyle(
            "fp-warn", parent=sample["BodyText"], fontName=FONT_REGULAR,
            fontSize=8.2, leading=11.5, textColor=colors.HexColor("#B3261E"),
        ),
        "cell": ParagraphStyle(
            "fp-cell", fontName=FONT_REGULAR, fontSize=7.4, leading=9.4,
        ),
    }


def report_to_pdf(report: Report) -> bytes:
    """The report as a document meant to be attached to somebody else's filing.

    Landscape, because the table is wide and shrinking it to portrait would either drop
    columns or set them at a size nobody reads.
    """
    styles = _pdf_styles()
    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=14 * mm, rightMargin=14 * mm, topMargin=13 * mm, bottomMargin=13 * mm,
        title=TITLE, author="FreightPrint",
    )

    story = [
        Paragraph(TITLE, styles["title"]),
        Paragraph(
            f"{report.factor_set} · {report.scope} · {date.today().isoformat()}",
            styles["body"],
        ),
        Spacer(1, 6 * mm),
    ]

    basis_rows = [
        [Paragraph(f"<b>{label}</b>", styles["cell"]), Paragraph(str(value), styles["cell"])]
        for label, value in _basis_lines(report)
    ]
    basis_rows.append([
        Paragraph("<b>Toplam CO2</b>", styles["cell"]),
        Paragraph(f"{_number_tr(round_to_significant(report.total_co2_kg))} kg", styles["cell"]),
    ])
    basis_rows.append([
        Paragraph("<b>Toplam fark</b>", styles["cell"]),
        Paragraph(f"{_number_tr(round_to_significant(report.total_saving_co2_kg))} kg", styles["cell"]),
    ])
    basis = Table(basis_rows, colWidths=[38 * mm, 225 * mm])
    basis.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LINEBELOW", (0, 0), (-1, -2), 0.25, colors.HexColor("#DDE2E9")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
    ]))
    story += [KeepTogether([Paragraph("<b>Hesap esası</b>", styles["body"]), Spacer(1, 2 * mm), basis])]

    if report.warnings:
        story += [Spacer(1, 5 * mm), Paragraph("<b>Uyarılar</b>", styles["body"])]
        story += [Paragraph(f"• {w}", styles["warn"]) for w in report.warnings]

    story += [Spacer(1, 6 * mm), Paragraph("<b>Sevkiyatlar</b>", styles["body"]), Spacer(1, 2 * mm)]

    head = [Paragraph(f"<b>{HEADER_TR[c]}</b>", styles["cell"]) for c in PDF_COLUMNS]
    body = []
    for report_row in report.rows:
        values = _row_values(report_row)
        body.append([
            Paragraph(_cell_text_for(c, values[c]), styles["cell"]) for c in PDF_COLUMNS
        ])

    # Widths in millimetres and summing to the printable width (A4 landscape less the
    # margins, 269 mm). Given as bare numbers they would be points, which packs the whole
    # table into the left third of the page and breaks place names mid-word.
    table = Table([head] + body, repeatRows=1,
                  colWidths=[w * mm for w in (22, 34, 34, 10, 42, 20, 24, 24, 20, 38)])
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B5FB0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#DDE2E9")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6F8FA")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]
    table.setStyle(TableStyle(style))
    story.append(table)

    story += [PageBreak(), Paragraph("<b>Yöntem</b>", styles["body"]), Spacer(1, 2 * mm),
              Paragraph(METHOD_NOTE, styles["body"])]

    document.build(story)
    return buffer.getvalue()
