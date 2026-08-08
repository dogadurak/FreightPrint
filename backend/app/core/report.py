"""Turn a file of shipments into the carbon report a customer already has to produce.

The interviews behind this project put a bulk report ahead of any interactive tool: the
obligation to produce one already exists, and it is produced by hand today.
"""

import csv
import io
import re
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass, field
from typing import Callable

from .emissions import (
    DEFAULT_FACTOR_SET,
    DEFAULT_SCOPE,
    calculate_shipment,
    load_emission_factors,
    load_tree_factors,
    lowest_emission_first,
    tree_equivalent,
)
from .road import RoadRoutingError
from .route import find_route_alternatives
from .uncertainty import round_to_significant

REQUIRED_COLUMNS = ("origin_lon", "origin_lat", "destination_lon", "destination_lat")
OPTIONAL_COLUMNS = ("reference", "origin_name", "destination_name", "tonnage")

# An .xlsx is a zip, and every zip starts here. Sniffed rather than trusted from the
# extension: a renamed file should still work, and a mislabelled one should say why.
ZIP_MAGIC = b"PK\x03\x04"

# A spreadsheet expands far beyond its compressed size, so the upload cap that guards
# the CSV path does not guard this one. Cells, not bytes, are what cost us here.
MAX_WORKBOOK_CELLS = 200_000

OUTPUT_COLUMNS = (
    "reference",
    "origin_name",
    "destination_name",
    "tonnage",
    "route",
    "road_km",
    "sea_km",
    "rail_km",
    "total_km",
    "road_co2_kg",
    "sea_co2_kg",
    "rail_co2_kg",
    "total_co2_kg",
    "all_road_co2_kg",
    "saving_co2_kg",
    "trees_equivalent",
    "status",
)


class ReportInputError(ValueError):
    pass


@dataclass
class ShipmentRow:
    reference: str
    origin: tuple[float, float]
    destination: tuple[float, float]
    origin_name: str
    destination_name: str
    tonnage: float


@dataclass
class ReportRow:
    shipment: ShipmentRow
    status: str
    route_label: str = ""
    distance_by_mode: dict[str, float] = field(default_factory=dict)
    co2_by_mode: dict[str, float] = field(default_factory=dict)
    total_km: float = 0.0
    total_co2_kg: float = 0.0
    all_road_co2_kg: float | None = None
    saving_co2_kg: float | None = None
    trees: float | None = None


@dataclass
class Report:
    rows: list[ReportRow]
    factor_set: str
    scope: str
    sources: list[str]
    warnings: list[str]

    @property
    def calculated(self) -> list[ReportRow]:
        return [row for row in self.rows if row.status == "ok"]

    @property
    def failed(self) -> list[ReportRow]:
        return [row for row in self.rows if row.status != "ok"]

    @property
    def total_co2_kg(self) -> float:
        return sum(row.total_co2_kg for row in self.calculated)

    @property
    def total_saving_co2_kg(self) -> float:
        return sum(row.saving_co2_kg or 0.0 for row in self.calculated)


def _number(value: str, column: str, row_number: int, comma_is_decimal: bool) -> float:
    """Read a number the way this particular file writes it.

    A Turkish or continental Excel writes 29,4306 for what this code wants as 29.4306,
    and refusing those files would refuse most of the real ones. But `1,234` is a
    thousand to an English Excel and one-and-a-bit to a Turkish one, and guessing wrong
    on a tonnage is a silent factor of a thousand that no range check would catch.

    So the ambiguity is settled by evidence rather than by the shape of the number: a
    file that separates its columns with semicolons or tabs has freed the comma to be a
    decimal mark, and a comma-separated file has not. `comma_is_decimal` carries that
    finding down from the delimiter, so no single value has to be guessed at.
    """
    text = str(value).strip()
    try:
        return float(text)
    except (TypeError, ValueError):
        pass
    if comma_is_decimal and re.fullmatch(r"-?\d+,\d+", text):
        return float(text.replace(",", "."))
    raise ReportInputError(f"row {row_number}: {column} is not a number: {text!r}")


def _cell_text(value) -> str:
    """A spreadsheet cell as the text an equivalent CSV would have held."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        # Excel stores every number as a float, so a tonnage of 24 arrives as 24.0
        # and a reference of 1001 as 1001.0. Neither should reach the report that way.
        return str(int(value))
    return str(value).strip()


def workbook_to_csv(raw: bytes) -> str:
    """Flatten the first worksheet into the CSV text the parser already understands.

    Only the first sheet is read. A workbook whose shipments are spread over several
    sheets would be silently half-reported, so this is a documented limit rather than
    a guess at which sheets were meant.
    """
    try:
        import openpyxl
    except ImportError as error:  # pragma: no cover - dependency is declared
        raise ReportInputError(
            "reading .xlsx needs openpyxl; install it or upload the sheet as CSV"
        ) from error

    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(raw), read_only=True, data_only=True, keep_links=False
        )
    except Exception as error:
        raise ReportInputError(f"could not read the workbook: {error}") from error

    try:
        sheet = workbook.worksheets[0]
        if sheet.max_row and sheet.max_column:
            if sheet.max_row * sheet.max_column > MAX_WORKBOOK_CELLS:
                raise ReportInputError(
                    f"sheet has more than {MAX_WORKBOOK_CELLS:,} cells; split it or upload CSV"
                )

        buffer = io.StringIO()
        writer = csv.writer(buffer, lineterminator="\n")
        written = 0
        for row in sheet.iter_rows(values_only=True):
            cells = [_cell_text(cell) for cell in row]
            if not any(cells):
                continue  # Trailing blank rows are Excel's, not the user's.
            writer.writerow(cells)
            written += 1
    finally:
        workbook.close()

    if not written:
        raise ReportInputError("the first worksheet is empty")
    return buffer.getvalue()


def read_upload(raw: bytes, filename: str | None = None) -> str:
    """Turn an uploaded file into CSV text, whichever of the two forms it arrived in."""
    if raw.startswith(ZIP_MAGIC):
        return workbook_to_csv(raw)
    if filename and filename.lower().endswith((".xlsx", ".xlsm")):
        raise ReportInputError(
            f"{filename} is named like a workbook but is not one. "
            "An .xls saved by an old Excel has to be re-saved as .xlsx or CSV."
        )
    try:
        return raw.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise ReportInputError(f"file must be UTF-8 text or an .xlsx workbook: {error}") from error


def parse_shipments(content: str) -> list[ShipmentRow]:
    """Read the upload, refusing it whole rather than reporting on a misread column."""
    # Excel in most of Europe writes CSV with semicolons, because the comma is its
    # decimal mark. Sniffed from the header alone, which is the one line whose shape
    # is known in advance.
    header = content.split("\n", 1)[0]
    delimiter = max(",;\t", key=header.count) if header else ","
    # Only a file that gave up the comma as a separator can be using it as a decimal mark.
    comma_is_decimal = delimiter != ","
    reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
    if reader.fieldnames is None:
        raise ReportInputError("file is empty")

    missing = [column for column in REQUIRED_COLUMNS if column not in reader.fieldnames]
    if missing:
        raise ReportInputError(
            f"missing column(s): {', '.join(missing)}. "
            f"Required: {', '.join(REQUIRED_COLUMNS)}; optional: {', '.join(OPTIONAL_COLUMNS)}"
        )

    shipments = []
    for number, row in enumerate(reader, start=2):
        read = lambda column: _number(row[column], column, number, comma_is_decimal)
        origin = (read("origin_lon"), read("origin_lat"))
        destination = (read("destination_lon"), read("destination_lat"))
        raw_tonnage = (row.get("tonnage") or "").strip() if row.get("tonnage") else ""
        tonnage = read("tonnage") if raw_tonnage else 24.0

        for point in (origin, destination):
            if not (-180 <= point[0] <= 180 and -90 <= point[1] <= 90):
                raise ReportInputError(f"row {number}: {point} is not a valid lon,lat pair")
        if tonnage <= 0:
            raise ReportInputError(f"row {number}: tonnage must be positive, got {tonnage}")

        shipments.append(
            ShipmentRow(
                reference=(row.get("reference") or f"row-{number}").strip(),
                origin=origin,
                destination=destination,
                origin_name=(row.get("origin_name") or "origin").strip(),
                destination_name=(row.get("destination_name") or "destination").strip(),
                tonnage=tonnage,
            )
        )

    if not shipments:
        raise ReportInputError("file has a header but no shipments")
    return shipments


def build_report(
    shipments: list[ShipmentRow],
    scope: str = DEFAULT_SCOPE,
    factor_set: str = DEFAULT_FACTOR_SET,
    road_fuel_type: str | None = None,
    load_factor: float | None = None,
    empty_return_share: float | None = None,
    concurrency: int = 1,
    on_progress: Callable[[int], None] | None = None,
) -> Report:
    """Price every shipment, keeping going when one cannot be routed.

    A single unroutable address must not cost the customer the other 33 rows, so a
    failure is recorded against its own row and the run continues.

    Rows are independent, so they can be routed a few at a time. Concurrency stays the
    caller's choice because the limit is the routing server's patience, not ours.
    """
    factors = load_emission_factors()
    tree_factors = load_tree_factors()
    warnings: set[str] = set()

    def price(shipment: ShipmentRow) -> ReportRow:
        try:
            routes = find_route_alternatives(
                origin=shipment.origin,
                destination=shipment.destination,
                origin_name=shipment.origin_name,
                destination_name=shipment.destination_name,
            )
            priced = calculate_shipment(
                routes,
                tonnage=shipment.tonnage,
                scope=scope,
                road_fuel_type=road_fuel_type,
                factor_set=factor_set,
                load_factor=load_factor,
                empty_return_share=empty_return_share,
            )
        except (RoadRoutingError, LookupError, ValueError) as error:
            return ReportRow(shipment=shipment, status=f"failed: {error}")

        ranked = lowest_emission_first(routes, priced)
        warnings.update(w for _, emission in ranked for w in emission.warnings)

        # The chosen row is the lowest-emission option, which is the question the report
        # answers. It may be the all-road baseline when nothing else beats it.
        route, emission = min(ranked, key=lambda pair: pair[1].total_co2_kg)
        saving = emission.saving_co2_kg
        return (
            ReportRow(
                shipment=shipment,
                status="ok",
                route_label=emission.label,
                distance_by_mode=route.distance_by_mode,
                co2_by_mode=emission.co2_by_mode,
                total_km=route.total_distance_km,
                total_co2_kg=emission.total_co2_kg,
                all_road_co2_kg=emission.all_road_co2_kg,
                saving_co2_kg=saving,
                trees=(
                    tree_equivalent(saving, tree_factors)["average_tree"]
                    if saving is not None
                    else None
                ),
            )
        )

    if concurrency > 1:
        with ThreadPoolExecutor(max_workers=concurrency) as pool:
            rows = []
            for row in pool.map(price, shipments):
                rows.append(row)
                if on_progress:
                    on_progress(len(rows))
    else:
        rows = []
        for shipment in shipments:
            rows.append(price(shipment))
            if on_progress:
                on_progress(len(rows))

    sources = sorted(
        {f.source for f in factors if f.factor_set == factor_set and f.scope == scope and f.is_verified}
    )
    return Report(
        rows=rows, factor_set=factor_set, scope=scope, sources=sources, warnings=sorted(warnings)
    )


def report_to_csv(report: Report) -> str:
    """Write the report, with the factor set and its sources in the file itself.

    A carbon figure without the basis it was produced on cannot be checked by whoever
    receives it, so the provenance travels with the numbers.
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")

    writer.writerow([f"# FreightPrint report | factor set: {report.factor_set} | scope: {report.scope}"])
    writer.writerow([f"# sources: {'; '.join(report.sources) or 'UNVERIFIED'}"])
    for warning in report.warnings:
        writer.writerow([f"# warning: {warning}"])
    writer.writerow([])
    writer.writerow(OUTPUT_COLUMNS)

    for row in report.rows:
        writer.writerow(
            [
                row.shipment.reference,
                row.shipment.origin_name,
                row.shipment.destination_name,
                row.shipment.tonnage,
                row.route_label,
                round(row.distance_by_mode.get("road", 0.0), 1),
                round(row.distance_by_mode.get("sea", 0.0), 1),
                round(row.distance_by_mode.get("rail", 0.0), 1),
                round(row.total_km, 1),
                round_to_significant(row.co2_by_mode.get("road", 0.0)),
                round_to_significant(row.co2_by_mode.get("sea", 0.0)),
                round_to_significant(row.co2_by_mode.get("rail", 0.0)),
                round_to_significant(row.total_co2_kg),
                round_to_significant(row.all_road_co2_kg) if row.all_road_co2_kg else "",
                round_to_significant(row.saving_co2_kg) if row.saving_co2_kg is not None else "",
                round(row.trees) if row.trees is not None else "",
                row.status,
            ]
        )

    writer.writerow([])
    writer.writerow(["# shipments calculated", len(report.calculated)])
    writer.writerow(["# shipments failed", len(report.failed)])
    writer.writerow(["# total CO2 kg", round_to_significant(report.total_co2_kg)])
    writer.writerow(["# total saving kg", round_to_significant(report.total_saving_co2_kg)])
    return buffer.getvalue()
